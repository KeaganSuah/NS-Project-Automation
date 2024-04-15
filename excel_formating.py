import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
import datetime
from parade_state_function import name_in_branch_generator, sort, ranking, amended_particulars_generator, branches_string
from copy import copy

# status used for excel
convert_status = {"Present": 'P', "Work From Home": 'WFH',
                  "Outside Stationed": 'OS',
                  "Attached Out": 'ATT',
                  "On Course": 'C', "Day Off": 'OFF', "Local Leave": 'LL',
                  "Overseas Leave": 'OL', "Medical Leave": 'MC',
                  "Medical / Dental appointment": 'MA',
                  "Report Sick Inside": 'RSI', "Report Sick Outside": 'RSO',
                  "Hospitalised / Sickbay": 'H', "AWOL": 'AWOL', "OTHERS": 'OTHERS'}


# format the type of border
def border_type(border):
    chosen_border = Border(left=Side(style=border),
                           right=Side(style=border),
                           top=Side(style=border),
                           bottom=Side(style=border))
    return chosen_border


# parade state Excel file, as well as different branches sheets
wb = openpyxl.load_workbook("website/PARADE_STATE.xlsx")

ws3 = wb['PARADE STATE S3']
ws3s = wb['SUMMARY S3']
ws1 = wb['PARADE STATE S1']
ws1s = wb['SUMMARY S1']
ws4 = wb['PARADE STATE S4']
ws4s = wb['SUMMARY S4']
ws2 = wb['PARADE STATE S2']
ws2s = wb['SUMMARY S2']
ws5 = wb['PARADE STATE S5']
ws5s = wb['SUMMARY S5']

type_xl_ps = [ws1, ws2, ws3, ws4, ws5]
type_xl_summary = [ws1s, ws2s, ws3s, ws4s, ws5s]


# used to track the number of personnel in Excel sheet
def value_counter(worksheet):
    value_record = 0
    for cells in worksheet.iter_rows(min_row=1,
                                     min_col=3, max_col=3,
                                     values_only=True):
        value = cells[0]
        if value is not None:
            value_record += 1
    return value_record


# multiple functions needed to run at the start of each month
def excel_monthly_reset(sheet):
    dt = datetime.datetime.today()
    SG_time = dt + datetime.timedelta(hours=17)
    day = SG_time.day
    if day == 1:
        clear_excel_range(sheet, 4, 35, 4, (value_counter(sheet) * 2) + 2)
        place_borders(sheet, 4, 36, 4, (value_counter(sheet) * 2) + 2)
        sort_by_rank(sheet)


# transfer data from the website into the Excel file, focus on parade state
def transfer_attendance_excel(dictionary, sheet):
    dt = datetime.datetime.today()
    SG_time = dt + datetime.timedelta(hours=17)
    day = SG_time.day
    excel_monthly_reset(sheet)
    for number in range(4, (value_counter(sheet) * 2) + 2):
        name = sheet.cell(row=number, column=3).value
        name = str(name).lower()
        for no, person in dictionary.items():
            if name == no.lower():
                sheet.cell(row=number, column=(day + 3)).value = str(convert_status[person[2]])
                sheet.cell(row=number + 1, column=(day + 3)).value = str(
                    convert_status[person[3]])
            else:
                continue
    change_date(sheet)


# used to clear a range of cells
def clear_excel_range(sheet, r_top_left, c_top_right, c_bottom_left, r_bottom_right):
    for row in range(r_top_left, r_bottom_right):
        for column in range(c_bottom_left, c_top_right):
            sheet.cell(row=row, column=column).value = None


# place borders for a range of cells
def place_borders(sheet, r_top_left, c_top_right, c_bottom_left, r_bottom_right):
    for row in range(r_top_left, r_bottom_right):
        for column in range(c_bottom_left, c_top_right):
            sheet.cell(row=row, column=column).border = border_type('thin')
            sheet.cell(row=row, column=column).fill = openpyxl.styles.PatternFill(fill_type=None)


def change_date(sheet):
    dt = datetime.datetime.today()
    SG_time = dt + datetime.timedelta(hours=17)
    year = SG_time.year
    month = SG_time.month
    day = SG_time.day
    sheet.cell(row=1, column=2).value = str(day) + "-" + str(month) + "-" + str(year)[2:4]


def save_file():
    wb.save("website/PARADE_STATE.xlsx")


# transfer data from the website into the Excel file, focus on summary
def transfer_summary_excel(dictionary, branch, sheet):
    for row in range(2, 21):
        for column in range(3, 22):
            status = sheet.cell(row=row, column=2).value
            status = str(status).lower()
            rank = sheet.cell(row=2, column=column).value
            rank = str(rank).lower()
            for a, b in dictionary.items():
                for name in b:
                    if status == a.lower() and rank == name.lower():
                        amount = sheet.cell(row=row, column=column).value
                        if amount is not None:
                            amount = int(amount)
                            amount += 1
                            sheet.cell(row=row, column=column).value = amount
                        else:
                            sheet.cell(row=row, column=column).value = 1
    for column in range(3, 23):
        rank = sheet.cell(row=2, column=column).value
        rank = str(rank).lower()
        for ranks in branch:
            if str(ranks) == rank:
                amount = sheet.cell(row=3, column=column).value
                if amount is not None:
                    amount = int(amount)
                    amount += 1
                    sheet.cell(row=3, column=column).value = amount
                else:
                    sheet.cell(row=3, column=column).value = 1


# insert new personnel at the bottom of the list in file
def insert_new_personnel(sheet, rank, name):
    sheet.insert_rows(idx=((value_counter(sheet) * 2) + 2), amount=2)
    for column in range(1, 4):
        sheet.merge_cells(start_row=((value_counter(sheet) * 2) + 2), start_column=column,
                          end_row=((value_counter(sheet) * 2) + 3), end_column=column)
    for row in range(((value_counter(sheet) * 2) + 2), ((value_counter(sheet) * 2) + 4)):
        for column in range(1, 36):
            sheet.cell(row=row, column=column).border = border_type('thin')
            sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=((value_counter(sheet) * 2) + 2), column=3).fill = PatternFill(start_color='00FFFF00',
                                                                                  end_color='00FFFF00',
                                                                                  fill_type='solid')
    sheet.cell(row=((value_counter(sheet) * 2) + 2), column=1).value = "=row()/2-1"
    sheet.cell(row=((value_counter(sheet) * 2) + 2), column=2).value = str(rank.upper())
    sheet.cell(row=((value_counter(sheet) * 2) + 2), column=3).value = str(name.upper())


# function tracks weekend date and red out the entire column
def block_weekend_excel(sheet):
    present_date = datetime.datetime.today()
    SG_time = present_date + datetime.timedelta(hours=17)
    present_date_day = SG_time.strftime('%d')
    first_day_month = SG_time - datetime.timedelta(days=int(present_date_day) - 1)
    x = first_day_month.weekday()
    for current_day in range(1, 32):
        if x == 5:
            for row in range(2, (value_counter(sheet))+1):
                sheet.cell(row=row * 2, column=current_day + 3).border = border_type(None)
                sheet.cell(row=row * 2, column=current_day + 3).fill = PatternFill(start_color='00FF0000',
                                                                                   end_color='00FF0000',
                                                                                   fill_type='solid')
            x += 1
        elif x == 6:
            for row in range(2, (value_counter(sheet))+1):
                sheet.cell(row=row * 2, column=current_day + 3).border = border_type(None)
                sheet.cell(row=row * 2, column=current_day + 3).fill = PatternFill(start_color='00FF0000',
                                                                                   end_color='00FF0000',
                                                                                   fill_type='solid')
            x = 0
        else:
            x += 1


# function that manage and track removed personnel
def tracking_delete(name, depot):
    count = 0
    for string in branches_string:
        if str(depot).lower() == str(string).lower():
            for number in range(2, (value_counter(type_xl_ps[count])) + 1):
                name_in_file = type_xl_ps[count].cell(row=number * 2, column=3).value
                if str(name_in_file) is not None:
                    if str(name_in_file).lower() == str(name).lower():
                        for column in range(1, 4):
                            type_xl_ps[count].unmerge_cells(start_row=(value_counter(type_xl_ps[count]) * 2),
                                                            start_column=column,
                                                            end_row=((value_counter(type_xl_ps[count]) * 2) + 1),
                                                            end_column=column)
                        type_xl_ps[count].delete_rows(idx=number * 2, amount=2)
        else:
            count += 1
            continue


unsorted_ranks = dict()
sorted_ranks = dict()


# dictionaries above and function below used to sort parade state sheet according to rank
def sort_by_rank(sheet):
    for number in range(2, (value_counter(sheet))+1):
        name = sheet.cell(row=number * 2, column=3).value
        name = str(name).lower()
        rank = sheet.cell(row=number * 2, column=2).value
        rank = str(rank).lower()
        unsorted_ranks[name] = [rank]
    sort(sorted_ranks, unsorted_ranks, ranking, 0)
    replace_value(sheet, sorted_ranks)


# same as the amended_particulars function, used to change attributes of personnel in Excel sheet
def amended_particulars_excel(sheet, search_name, input_rank, input_name):
    # if input name is filled, update new name and rank
    if input_name != '':
        for number in range(2, (value_counter(sheet))+1):
            name = str(sheet.cell(row=number * 2, column=3).value).lower()
            if name == search_name.lower():
                sheet.cell(row=number * 2, column=2).value = str(input_rank.upper())
                sheet.cell(row=number * 2, column=3).value = str(input_name.upper())
        amended_particulars_generator(input_rank, search_name, input_name)
        save_file()
    # if input name is empty, update rank only
    else:
        for number in range(2, (value_counter(sheet))+1):
            name = sheet.cell(row=number * 2, column=3).value
            if name == search_name.upper():
                sheet.cell(row=number * 2, column=2).value = str(input_rank.upper())
                amended_particulars_generator(input_rank, search_name, input_name)
        save_file()


def replace_value(sheet, dictionary):
    number = 4
    for name, rank in dictionary.items():
        sheet.cell(row=number, column=2).value = str(rank[0].upper())
        sheet.cell(row=number, column=3).value = str(name.upper())
        number += 2
    dictionary.clear()


def transfer_master_depots(master, depot):
    for row in range(1, value_counter(depot)):
        for column in range(1, 4):
            depot.unmerge_cells(start_row=((row * 2) + 2), start_column=column, end_row=((row * 2) + 3),
                                end_column=column)
    depot.delete_rows(idx=1, amount=(value_counter(depot) * 2) + 30)
    for row in master.rows:
        for cell in row:
            new_cell = depot.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
    for row in range(1, value_counter(depot)):
        for column in range(1, 4):
            depot.merge_cells(start_row=((row * 2) + 2), start_column=column, end_row=((row * 2) + 3),
                              end_column=column)


def transfer_master_depot_summary(master, depot):
    mr = master.max_row
    mc = master.max_column
    for i in range(2, mr + 1):
        for j in range(1, mc + 1):
            c = master.cell(row=i, column=j)
            depot.cell(row=i, column=j).value = c.value
