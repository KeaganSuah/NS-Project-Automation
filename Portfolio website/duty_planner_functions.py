import datetime
import openpyxl
import calendar
from tinydb import TinyDB, Query

User = Query()
DO_db = TinyDB('website/DO_db.json')

wb = openpyxl.load_workbook("website/Duty Roster.xlsx")
DO_FORECAST = wb['DO FORECAST']
DO_POINTS = wb['DO POINTS']

unavailability_reasons = {"Exercise Support": [], "Surgery": [], "On Course": [], "Duty Rest": [], "Birthday": [],
                          "Local Leave": [],
                          "Conducting Body": [], "Medical Appointment": [], "Outstation": [],
                          "Day Off": [], "study commitment": [], "personnel commitment": []}
namelist = {}


def block_out_month(block_list):
    for name, particulars in namelist.copy().items():
        for reason, block_dates in particulars['unavailable'].items():
            if reason in list(unavailability_reasons.keys())[:3]:
                if len(block_dates) > 23:
                    block_list.append(str(name))
                    particulars['unavailable'][str(reason)] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16,
                                                               17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29,
                                                               30, 31]
                else:
                    continue
            else:
                continue


# function to enter details into the database
def duty_insert(name, month, points, cooldown, cooldown_s, excuse, depot, database):
    database.insert(
        {'name': name, 'month': month, 'points': points, 'cooldown': cooldown, 'cooldown_s': cooldown_s,
         'excuse': excuse, 'depot': depot,
         'unavailable': {'Exercise Support': [], 'Surgery': [], 'On Course': [], 'Duty Rest': [], 'Birthday': [],
                         'Local Leave': [],
                         'Conducting Body': [], 'Medical Appointment': [], 'Outstation': [],
                         'Day Off': [],
                         'study commitment': [], 'personnel commitment': []},
         'available': []})


# function to remove a personnel details from the database
def duty_delete_name(name, database):
    database.remove(User.name == name)


# function to amend a personnel details from the database
def duty_amend_name_depot(input_rank, search_name, input_name, depot, database):
    database.update({'depot': str(depot)}, User.name == search_name)
    database.update({'name': str(input_rank + ' ' + input_name)}, User.name == search_name)


def duty_add_point(database, name, point):
    database.update({'points': (database.search(User.name == name.upper())[0]['points'] + int(point))},
                    User.name == name.upper())


# function that extract the updated database into the duty planner function
def extract_database(database):
    duty_database = database.all()
    for duty_personnel in duty_database:
        namelist[duty_personnel.get('name')] = {'month': duty_personnel.get('month'),
                                                'points': duty_personnel.get('points'),
                                                'cooldown': duty_personnel.get('cooldown'),
                                                'cooldown_s': duty_personnel.get('cooldown_s'),
                                                'excuse': duty_personnel.get('excuse'),
                                                'depot': duty_personnel.get('depot'),
                                                'unavailable': duty_personnel.get('unavailable'),
                                                'available': duty_personnel.get('available')}


weekend_days = []


def block_weekend_excel(defer_time):
    weekend_days.clear()
    present_date = datetime.datetime.today()
    SG_time = present_date + datetime.timedelta(days=defer_time)
    present_date_day = SG_time.strftime('%d')
    first_day_month = SG_time - datetime.timedelta(days=int(present_date_day) - 1)
    x = first_day_month.weekday()
    for current_day in range(1, 32):
        if x == 5:
            weekend_days.append(current_day)
            x += 1
        elif x == 6:
            weekend_days.append(current_day)
            x = 0
        else:
            x += 1


days_in_month = dict()


def duty_rouster_generator(time):
    days_in_month.clear()
    present_date = datetime.datetime.today()
    SG_time = present_date + datetime.timedelta(days=time)
    daysInMonth = calendar.monthrange(SG_time.year, SG_time.month)[1]
    for days in range(1, daysInMonth + 1):
        days_in_month[str(days)] = {'duty': [], 'standby': ''}
    block_weekend_excel(time)


def generate_average(dictionary):
    for name, details in dictionary.items():
        details['average'] = details['points'] / details['month']


normal_duty = {}
excuse_duty = {}


def separate_excuse_normal():
    generate_average(namelist)
    for name, details in namelist.items():
        if details['excuse'] == 'N':
            normal_duty[name] = details
            details.pop('excuse')
        else:
            excuse_duty[name] = details
            details.pop('excuse')


def priority_sorter(dictionary):
    for sorted_name in sorted(dictionary, key=lambda unavailable: dictionary[unavailable]['average'],
                              reverse=False):
        for name, details in dictionary.copy().items():
            if sorted_name == name:
                dictionary.pop(sorted_name)
                dictionary[sorted_name] = details
            else:
                continue


def reverse_priority_sorter(dictionary):
    for sorted_name in sorted(dictionary, key=lambda unavailable: dictionary[unavailable]['average'],
                              reverse=True):
        for name, details in dictionary.copy().items():
            if sorted_name == name:
                dictionary.pop(sorted_name)
                dictionary[sorted_name] = details
            else:
                continue


sign_extra_do = {}
sign_extra_doo = {}


def sign_extra_generator(dictionary):
    for name, sign_extra_days in dictionary.copy().items():
        sign_extra = [int(item) for item in sign_extra_days.split(',')]
        for days in sign_extra:
            days_in_month[str(days)]['duty'].append(str(name))
            days_in_month[str(days)]['duty'].append(str(name))
        namelist[str(name)]['cooldown'] = sign_extra_days[-1] + 7
        dictionary.pop(name)


def average_of_normal():
    generate_average(namelist)
    x = 0
    y = 0
    for name, details in namelist.copy().items():
        x += 1
        y += int(details['average'])
    return y / x


def duty_plan_generator(cool_days, cool_days_s, database, sheet_forecast, sheet_points, sign_extra_dictionary,
                        duty_string, empty, clearing_dictionary):
    for lists in empty:
        lists.clear()
    block_out_list = []
    personnel_doing_weekend = []
    namelist.clear()
    normal_duty.clear()
    excuse_duty.clear()
    clearing_dictionary.clear()
    extract_database(database)
    database.purge()
    duty_points_excel(database, sheet_points)
    block_out_month(block_out_list)
    duty_rouster_generator(10)
    sign_extra_generator(sign_extra_dictionary)
    separate_excuse_normal()
    generate_average(normal_duty)
    reverse_priority_sorter(normal_duty)
    for name, details in normal_duty.copy().items():
        count = 0
        for day, duty in days_in_month.items():
            if len(duty['duty']) == 0:
                if int(day) in details['available']:
                    if count > 1:
                        continue
                    else:
                        unavailable_list = []
                        for lists in list(details['unavailable'].values()):
                            unavailable_list += lists
                        if int(day) in unavailable_list:
                            continue
                        else:
                            duty['duty'].append(name)
                            normal_duty[name]['unavailable']['Duty Rest'].extend(range(int(day) - 8, int(day) + 8 + 1))
                            if int(day) in weekend_days:
                                details['points'] += 3
                                count += 2
                                personnel_doing_weekend.append(name)
                            else:
                                details['points'] += 1
                                count += 1
                                normal_duty[name]['unavailable']['Duty Rest'].extend(weekend_days)
                                if count > 1:
                                    details['cooldown'] = (cool_days * 2) + 2
                else:
                    continue
            else:
                continue
    generate_average(normal_duty)
    priority_sorter(normal_duty)
    for first_weekend_day in weekend_days[0:2]:
        for name, details in normal_duty.copy().items():
            if len(days_in_month[str(first_weekend_day)]['duty']) == 0:
                unavailable_list = []
                for lists in list(details['unavailable'].values()):
                    unavailable_list += lists
                if int(first_weekend_day) in unavailable_list:
                    continue
                else:
                    if details['cooldown'] > 7:
                        continue
                    else:
                        days_in_month[str(first_weekend_day)]['duty'].append(name)
                        details['points'] += 3
                        details['cooldown'] = weekend_days[1] + cool_days
                        normal_duty.pop(name)
                        normal_duty[name] = details
                        personnel_doing_weekend.append(name)
    for second_weekend_day in weekend_days[2:]:
        for name, details in normal_duty.copy().items():
            if len(days_in_month[str(second_weekend_day)]['duty']) == 0:
                unavailable_list = []
                for lists in list(details['unavailable'].values()):
                    unavailable_list += lists
                if int(second_weekend_day) in unavailable_list:
                    continue
                else:
                    days_in_month[str(second_weekend_day)]['duty'].append(name)
                    details['points'] += 3
                    normal_duty[name]['unavailable']['Duty Rest'].extend(
                        range(int(second_weekend_day) - (cool_days - 7), int(second_weekend_day) + (cool_days - 7) + 1))
                    normal_duty.pop(name)
                    normal_duty[name] = details
                    personnel_doing_weekend.append(name)
            else:
                continue
    generate_average(normal_duty)
    priority_sorter(normal_duty)
    for day, duty in days_in_month.items():
        for name, details in normal_duty.copy().items():
            details['cooldown'] -= 1
            if len(duty['duty']) == 0:
                unavailable_list = []
                for lists in list(details['unavailable'].values()):
                    unavailable_list += lists
                if int(day) in unavailable_list:
                    continue
                else:
                    if details['cooldown'] > 0:
                        continue
                    else:
                        duty['duty'].append(name)
                        details['cooldown'] = cool_days
                        if int(day) in weekend_days:
                            details['points'] += 3
                            normal_duty.pop(name)
                            normal_duty[name] = details
                        else:
                            details['points'] += 1
                            normal_duty.pop(name)
                            normal_duty[name] = details
            else:
                continue
    for day, duty in days_in_month.items():
        for name in duty['duty']:
            normal_duty[name]['unavailable']['Duty Rest'].extend(range(int(day) - 6, int(day) + 6 + 1))
    for weekend_day in list(days_in_month.keys())[-7:]:
        name = days_in_month[str(weekend_day)]['duty'][0]
        if name is not None:
            normal_duty[str(name)]['cooldown'] = int(weekend_day) - int(list(days_in_month.keys())[-8])
    generate_average(normal_duty)
    priority_sorter(normal_duty)
    for day, duty in days_in_month.items():
        for name, details in normal_duty.copy().items():
            details['cooldown_s'] -= 1
            if duty['standby'] == '':
                unavailable_list = []
                for lists in list(details['unavailable'].values()):
                    unavailable_list += lists
                if int(day) in unavailable_list:
                    continue
                else:
                    if details['cooldown_s'] > 0:
                        continue
                    else:
                        if str(sheet_forecast) == '<Worksheet "DO FORECAST">':
                            if namelist[duty['duty'][0]].get('depot') == details['depot']:
                                if int(day) in weekend_days and name in personnel_doing_weekend:
                                    continue
                                else:
                                    duty['standby'] = name
                                    details['cooldown_s'] = cool_days_s
                                    normal_duty.pop(name)
                                    normal_duty[name] = details
                                    if int(day) > int(list(days_in_month)[-3]):
                                        details['cooldown'] = int(day) - int(list(days_in_month)[-3])
                            else:
                                continue
                        else:
                            if int(day) in weekend_days and name in personnel_doing_weekend:
                                continue
                            else:
                                duty['standby'] = name
                                details['cooldown_s'] = cool_days_s
                                normal_duty.pop(name)
                                normal_duty[name] = details
                                if int(day) > int(list(days_in_month)[-1]):
                                    details['cooldown'] = int(day) - int(list(days_in_month)[-1])
            else:
                continue
    for i in range(12, 13):
        if str(sheet_forecast) == '<Worksheet "DO FORECAST">':
            fill_empty(i)
    check_if_empty(empty, 0)
    check_if_empty_standby(empty, 1)
    for i in reversed(range(9, 11)):
        fill_empty(i)
    for i in reversed(range(9, 11)):
        fill_empty_standby(i)
    check_if_empty(empty, 2)
    check_if_empty_standby(empty, 3)
    generate_average(excuse_duty)
    reverse_priority_sorter(excuse_duty)
    for name, details in excuse_duty.copy().items():
        count = 0
        for day, duty in days_in_month.items():
            if len(duty['duty']) > 1:
                if int(day) in details['available']:
                    if count > 1:
                        continue
                    else:
                        if int(day) in weekend_days:
                            unavailable_list = []
                            for lists in list(details['unavailable'].values()):
                                unavailable_list += lists
                            if int(day) in unavailable_list:
                                continue
                            else:
                                duty['duty'].append(name)
                                count += 1
                                excuse_duty[name]['unavailable']['Duty Rest'].extend(
                                    range(int(day) - 6, int(day) + 6 + 1))
                                normal_duty[duty['duty'][0]]['points'] -= 1.5
                                details['points'] += 1.5
                        else:
                            continue
                else:
                    continue
            else:
                continue
    generate_average(excuse_duty)
    priority_sorter(excuse_duty)
    for date, duties in days_in_month.items():
        for excuse_name, weekend_duty in excuse_duty.copy().items():
            weekend_duty['cooldown'] -= 1
            if int(date) in weekend_days:
                if len(duties['duty']) > 1:
                    continue
                else:
                    unavailable_list = []
                    for lists in list(weekend_duty['unavailable'].values()):
                        unavailable_list += lists
                    if int(date) in unavailable_list:
                        continue
                    else:
                        if weekend_duty['cooldown'] > 0:
                            continue
                        else:
                            normal_duty[duties['duty'][0]]['points'] -= 1.5
                            duties['duty'].append(excuse_name)
                            weekend_duty['cooldown'] = 10
                            weekend_duty['points'] += 1.5
                            excuse_duty.pop(excuse_name)
                            excuse_duty[excuse_name] = weekend_duty
            else:
                continue
    for name, details in normal_duty.copy().items():
        if str(name) in block_out_list:
            duty_insert(str(name).upper(), details['month'], details['points'], details['cooldown'],
                        details['cooldown_s'],
                        'N', details['depot'], database)
        else:
            duty_insert(str(name).upper(), details['month'] + 1, details['points'], details['cooldown'],
                        details['cooldown_s'],
                        'N', details['depot'], database)
    for name, details in excuse_duty.copy().items():
        if str(name) in block_out_list:
            duty_insert(str(name).upper(), details['month'], details['points'], details['cooldown'],
                        details['cooldown_s'],
                        'Y', details['depot'], database)
        else:
            duty_insert(str(name).upper(), details['month'] + 1, details['points'], details['cooldown'],
                        details['cooldown_s'],
                        'Y', details['depot'], database)
    duty_planned_excel(sheet_forecast, days_in_month)
    present_date = datetime.datetime.today()
    SG_time = present_date + datetime.timedelta(days=10)
    month_number = SG_time.month
    text = duty_string + ' DUTY ROUSTER FOR ' + str(calendar.month_name[month_number]).upper()
    sheet_forecast.cell(row=1, column=1).value = text
    save_file()


empty_do = [[], [], [], []]
empty_doo = [[], [], [], []]


def check_if_empty(empty_list, position):
    for day, duty in days_in_month.items():
        if len(duty['duty']) == 0:
            empty_list[position].append(str(day))


def check_if_empty_standby(empty_list, position):
    for day, duty in days_in_month.items():
        if duty['standby'] == '':
            empty_list[position].append(str(day))


def check_and_fill(day, duty, x, y, reason_position):
    if len(duty['duty']) == 0:
        generate_average(normal_duty)
        priority_sorter(normal_duty)
        for name, details in normal_duty.copy().items():
            name_in_list = []
            for block_date in range(x, y):
                name_in_list.append(str(days_in_month[str(block_date)]['standby']))
                name_in_list.extend(days_in_month[str(block_date)]['duty'])
            if name in name_in_list:
                continue
            else:
                unavailable_list = []
                for lists in list(details['unavailable'].values())[:reason_position]:
                    unavailable_list += lists
                if int(day) in unavailable_list:
                    continue
                else:
                    duty['duty'].append(name)
                    if int(day) in weekend_days:
                        details['points'] += 3
                        break
                    else:
                        details['points'] += 1
                        break


def check_and_fill_standby(day, duty, x, y, reason_position):
    if duty['standby'] == '':
        for name, details in normal_duty.copy().items():
            name_in_list = []
            for block_date in range(x, y):
                name_in_list.append(str(days_in_month[str(block_date)]['standby']))
                name_in_list.extend(days_in_month[str(block_date)]['duty'])
            if name in name_in_list:
                continue
            else:
                unavailable_list = []
                for lists in list(details['unavailable'].values())[:reason_position]:
                    unavailable_list += lists
                if int(day) in unavailable_list:
                    continue
                else:
                    duty['standby'] = name
                    normal_duty.pop(name)
                    normal_duty[name] = details


def fill_empty(reason_position):
    for day, duty in days_in_month.items():
        x = int(day) - 4
        y = int(day) + 4
        if x < 1:
            x = 1
            check_and_fill(day, duty, x, y, reason_position)
        elif y > int(list(days_in_month.keys())[-1]):
            y = int(list(days_in_month.keys())[-1])
            check_and_fill(day, duty, x, y, reason_position)
        else:
            check_and_fill(day, duty, x, y, reason_position)


def fill_empty_standby(reason_position):
    for day, duty in days_in_month.items():
        x = int(day) - 4
        y = int(day) + 4
        if x < 1:
            x = 1
            check_and_fill_standby(day, duty, x, y, reason_position)
        elif y > int(list(days_in_month.keys())[-1]):
            y = int(list(days_in_month.keys())[-1])
            check_and_fill_standby(day, duty, x, y, reason_position)
        else:
            check_and_fill_standby(day, duty, x, y, reason_position)


def points_swap(database, minus_points_name, plus_points_name, points):
    if minus_points_name.upper == 'BLANK' or len(database.search(User.name == minus_points_name.upper())) == 0:
        pass
    else:
        database.update({'points': (database.search(User.name == minus_points_name.upper())[0]['points'] - points)},
                        User.name == minus_points_name.upper())
    if len(database.search(User.name == plus_points_name.upper())) != 0:
        database.update({'points': (database.search(User.name == plus_points_name.upper())[0]['points'] + points)},
                        User.name == plus_points_name.upper())


def cooldown_swap(database, minus_cooldown_name, plus_cooldown_name, cooldown, cooldown_string):
    if minus_cooldown_name.upper == 'BLANK' or len(database.search(User.name == minus_cooldown_name.upper())) == 0:
        pass
    else:
        database.update({cooldown_string: 0}, User.name == minus_cooldown_name.upper())
    database.update({cooldown_string: cooldown}, User.name == str(plus_cooldown_name).upper())


def standby_activate(input_day, remove_name, database, sheet_current):
    current_duty_list = dict()
    excel_to_dictionary(sheet_current, current_duty_list)
    block_weekend_excel(0)
    current_duty_list[str(input_day)]['duty'].insert(current_duty_list[str(input_day)]['duty'].index(str(remove_name)),
                                                     (str(current_duty_list[str(input_day)]['standby'])))
    current_duty_list[str(input_day)]['duty'].remove(str(remove_name))
    if input_day in weekend_days:
        if len(current_duty_list[str(input_day)]['duty']) > 1:
            points_swap(database, remove_name.upper(), str(current_duty_list[str(input_day)]['standby']).upper(), 1.5)
        else:
            points_swap(database, remove_name.upper(), str(current_duty_list[str(input_day)]['standby']).upper(), 3)
    else:
        points_swap(database, remove_name.upper(), str(current_duty_list[str(input_day)]['standby']).upper(), 1)
    if int(input_day) > int(list(current_duty_list)[-8]):
        cooldown_swap(database, remove_name.upper(), str(current_duty_list[str(input_day)]['standby']).upper(),
                      int(input_day) - int(list(current_duty_list)[-12]), 'cooldown')
    duty_planned_excel(sheet_current, current_duty_list)
    save_file()


def swap_duty(first_day, first_name, second_day, second_name, database, sheet_current):
    current_duty_list = dict()
    excel_to_dictionary(sheet_current, current_duty_list)
    block_weekend_excel(0)
    current_duty_list[str(first_day)]['duty'].insert(current_duty_list[str(first_day)]['duty'].index(str(first_name)),
                                                     str(second_name))
    current_duty_list[str(second_day)]['duty'].insert(
        current_duty_list[str(second_day)]['duty'].index(str(second_name)), str(first_name))
    current_duty_list[str(first_day)]['duty'].remove(str(first_name))
    current_duty_list[str(second_day)]['duty'].remove(str(second_name))
    if int(first_day) in weekend_days:
        if len(current_duty_list[str(first_day)]['duty']) > 1:
            points_swap(database, first_name.upper(), second_name.upper(), 1.5)
        else:
            points_swap(database, first_name.upper(), second_name.upper(), 3)
    elif int(first_day) not in weekend_days:
        points_swap(database, first_name.upper(), second_name.upper(), 1)
    if int(second_day) in weekend_days:
        if len(current_duty_list[str(second_day)]['duty']) > 1:
            points_swap(database, second_name.upper(), first_name.upper(), 1.5)
        else:
            points_swap(database, second_name.upper(), first_name.upper(), 3)
    elif int(second_day) not in weekend_days:
        points_swap(database, second_name.upper(), first_name.upper(), 1)
    if int(first_day) > int(list(current_duty_list)[-8]):
        cooldown_swap(database, first_name.upper(), second_name.upper(),
                      int(first_day) - int(list(current_duty_list)[-12]), 'cooldown')
    if int(second_day) > int(list(current_duty_list)[-8]):
        cooldown_swap(database, second_name.upper(), first_name.upper(),
                      int(first_day) - int(list(current_duty_list)[-12]), 'cooldown')
    duty_planned_excel(sheet_current, current_duty_list)
    save_file()


def exchange_personnel(date, old_name, new_name, database, sheet_current):
    current_duty_list = dict()
    excel_to_dictionary(sheet_current, current_duty_list)
    block_weekend_excel(0)
    current_duty_list[str(date)]['duty'].insert(current_duty_list[str(date)]['duty'].index(str(old_name)),
                                                str(new_name))
    current_duty_list[str(date)]['duty'].remove(str(old_name))
    if int(date) in weekend_days:
        if len(current_duty_list[str(date)]['duty']) > 1:
            points_swap(database, old_name.upper(), new_name.upper(), 1.5)
        else:
            points_swap(database, old_name.upper(), new_name.upper(), 3)
    elif int(date) not in weekend_days:
        points_swap(database, old_name.upper(), new_name.upper(), 1)
    if int(date) > int(list(current_duty_list)[-8]):
        cooldown_swap(database, old_name.upper(), new_name.upper(),
                      int(date) - int(list(current_duty_list)[-12]), 'cooldown')
    duty_planned_excel(sheet_current, current_duty_list)
    save_file()


def exchange_personnel_standby(date, old_name, new_name, database, sheet_current):
    current_duty_list = dict()
    excel_to_dictionary(sheet_current, current_duty_list)
    block_weekend_excel(0)
    current_duty_list[str(date)]['standby'] = str(new_name)
    if int(date) > int(list(current_duty_list)[-8]):
        cooldown_swap(database, old_name.upper(), new_name.upper(),
                      int(date) - int(list(current_duty_list)[-8]), 'cooldown_s')
    duty_planned_excel(sheet_current, current_duty_list)
    save_file()


def save_file():
    wb.save("website/Duty Roster.xlsx")


def clear_excel_range(sheet, row1, row2, col1, col2):
    for row in range(row1, row2):
        for column in range(col1, col2):
            sheet.cell(row=row, column=column).value = None


def duty_points_excel(database, sheet):
    extract_database(database)
    clear_excel_range(sheet, 2, 100, 1, 4)
    clear_excel_range(sheet, 2, 100, 5, 20)
    row = 2
    for name, particulars in namelist.copy().items():
        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = particulars['month']
        sheet.cell(row=row, column=3).value = particulars['points']
        column = 5
        for unavailable, dates in particulars['unavailable'].items():
            if unavailable == 'Duty Rest':
                continue
            else:
                if len(dates) == 0:
                    sheet.cell(row=row, column=column).value = None
                else:
                    string = ', '.join(str(item) for item in dates)
                    sheet.cell(row=row, column=column).value = string
                column += 1
        if len(particulars['available']) == 0:
            sheet.cell(row=row, column=16).value = None
        else:
            string = ', '.join(str(item) for item in particulars['available'])
            sheet.cell(row=row, column=16).value = string
        row += 1
    save_file()


def duty_planned_excel(worksheet, dictionary):
    clear_excel_range(worksheet, 5, 36, 2, 7)
    for days, duty_n_standby in dictionary.copy().items():
        if len(duty_n_standby['duty']) == 2:
            y = 4
            for names in duty_n_standby['duty']:
                worksheet.cell(row=(int(days) + 4), column=y).value = str(names).upper()
                y -= 2
        elif len(duty_n_standby['duty']) == 1:
            worksheet.cell(row=(int(days) + 4), column=2).value = str(duty_n_standby['duty'][0]).upper()
        else:
            continue
        worksheet.cell(row=(int(days) + 4), column=6).value = str(duty_n_standby['standby']).upper()


def excel_to_dictionary(worksheet, dictionary):
    for days in range(1, 32):
        day = worksheet.cell(row=(int(days) + 4), column=1).value
        duty_name = worksheet.cell(row=(int(days) + 4), column=2).value
        duty_name2 = worksheet.cell(row=(int(days) + 4), column=4).value
        standby_name = worksheet.cell(row=(int(days) + 4), column=6).value
        if duty_name is None:
            if standby_name is None:
                dictionary[str(day)] = {'duty': ['BLANK'], 'standby': 'BLANK'}
            else:
                dictionary[str(day)] = {'duty': ['BLANK'], 'standby': standby_name}
        else:
            if duty_name2 is None:
                if standby_name is None:
                    dictionary[str(day)] = {'duty': [duty_name], 'standby': 'BLANK'}
                else:
                    dictionary[str(day)] = {'duty': [duty_name], 'standby': standby_name}
            else:
                if standby_name is None:
                    dictionary[str(day)] = {'duty': [duty_name2, duty_name], 'standby': 'BLANK'}
                else:
                    dictionary[str(day)] = {'duty': [duty_name2, duty_name], 'standby': standby_name}


def forecast_to_current(sheet_current, sheet_forecast, duty_string):
    current_duty_list = dict()
    present_date = datetime.datetime.today()
    SG_time = present_date + datetime.timedelta(days=5)
    month_number = SG_time.month
    text = duty_string + ' DUTY ROUSTER FOR ' + str(calendar.month_name[month_number]).upper()
    sheet_current.cell(row=1, column=1).value = text
    excel_to_dictionary(sheet_forecast, current_duty_list)
    duty_planned_excel(sheet_current, current_duty_list)
    save_file()


def generate_names_list(database, name_list):
    name_list.clear()
    for personnel in database.all():
        name_list.append(personnel.get('name'))


requested_blockout_doo = dict()
requested_blockout_do = dict()


def accepted_blockout(name, blockout_dictionary, database):
    dictionary = unavailability_reasons.copy()
    for excuse, dates in blockout_dictionary[name].items():
        days = [int(item) for item in dates.split(',')]
        days.sort()
        dictionary[excuse] = days
    database.update({'unavailable': dictionary}, User.name == name)


def submitted_available(database, name, dates):
    days = [int(item) for item in dates.split(',')]
    days.sort()
    database.update({'available': days}, User.name == name)


def generate_blockout_display(database, dictionary):
    dictionary.clear()
    for personnel in database.all():
        if personnel.get('name') not in dictionary.keys():
            if len(personnel.get('available')) > 0:
                dictionary[personnel.get('name')] = {}
                available_string = ', '.join(str(item) for item in personnel.get('available'))
                dictionary[personnel.get('name')]['Preferred Date'] = available_string
            else:
                continue
        else:
            if len(personnel.get('available')) > 0:
                available_string = ', '.join(str(item) for item in personnel.get('available'))
                dictionary[personnel.get('name')]['Preferred Date'] = available_string
            else:
                continue
    for personnel in database.all():
        for reason, dates in personnel.get('unavailable').items():
            if personnel.get('name') not in dictionary.keys():
                if len(dates) > 0:
                    string = ', '.join(str(item) for item in dates)
                    dictionary[personnel.get('name')] = {}
                    dictionary[personnel.get('name')][reason] = string
                else:
                    continue
            else:
                if len(dates) > 0:
                    string = ', '.join(str(item) for item in dates)
                    dictionary[personnel.get('name')][reason] = string
                else:
                    continue
